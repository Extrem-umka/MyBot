import tkinter as tk
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time
import threading

# Функция для выполнения бота
def run_bot():
    # Открываем или создаем файл Excel
    excel_file = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Excel Files", "*.xlsx")])
    if not excel_file:
        return
    workbook = openpyxl.load_workbook(excel_file)
    ws = workbook.active

    # Задания из файла в список
    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tasks.append((row[0], row[1]))

    # Запускаем хром
    driver = webdriver.Chrome()

    try:
        # Перебираем задания
        for i, task in enumerate(tasks):
            task_id, status = task

            # Заходим на страницу нашего id задания
            driver.get(f"{task_id}")

            time.sleep(1)

            try:
                return_button1 = driver.find_element(By.CLASS_NAME, "glyphEmpty")
                return_button1.click()
                return_button2 = driver.find_element(By.CLASS_NAME, "glyphEmpty")
                return_button2.click()
                return_button3 = driver.find_element(By.CLASS_NAME, "glyphEmpty")
                return_button3.click()

                ws.cell(row=i + 2, column=2).value = "успешно"
            except:
                ws.cell(row=i + 2, column=2).value = "ошибка"

            workbook.save(excel_file)
    finally:
        driver.quit()  # Закрываем окно браузера
        root.destroy()  # Закрываем окно tkinter

    # Запускаем бота в отдельном потоке
    bot_thread = threading.Thread(target=bot_thread)
    bot_thread.start()

# Функция для создания и сохранения шаблона Excel-файла
def create_template():
    template_workbook = openpyxl.Workbook()
    template_sheet = template_workbook.active
    template_sheet['A1'] = "Ссылка на обновление задания"
    template_sheet['B1'] = "Статус выполнения"
    template_workbook.save("RDB_tasks.xlsx")

# Создаем главное окно tkinter
root = tk.Tk()
root.title("Бот RDB для обновления задания")
root.geometry('350x50')
# Создаем кнопку для запуска бота
run_button = tk.Button(root, text="Выбрать шаблон и Запустить бота", command=run_bot)
run_button.pack()

# Создаем кнопку для создания и сохранения шаблона
template_button = tk.Button(root, text="Сохранить шаблон(RDB_tasks.xlsx)", command=create_template)
template_button.pack()

root.mainloop()