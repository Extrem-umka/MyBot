import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Открываем Excel файл

workbook = openpyxl.load_workbook('RDBTASKSFIND.xlsx')
sheet = workbook.active
try:
    # Цикл для обхода строк в Excel файле
    i = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        region, number, plan, character = row  # Распаковываем данные из строки

        # Запускаем браузер

        link = "https://***"
        browser = webdriver.Chrome()
        browser.get(link)

        try:
            button_create = browser.find_element(By.XPATH, "(//span[@id='p_body_ctl128'])[1]")
            button_create.click()
            time.sleep(3)

            iframe = browser.find_element(By.XPATH, "//iframe[@id='ewerywareModalFormIFrame']")
            browser.switch_to.frame(iframe)

            input_region = browser.find_element(By.XPATH,
                                                "(//input[@id='ctl00_body_mlkp_-1_lookupControl_lookupText'])[1]")
            input_region.send_keys(region)

            input_number = browser.find_element(By.XPATH, "(//input[@name='ctl00$body$ctl12'])[1]")
            input_number.send_keys(number)

            input_plan = browser.find_element(By.XPATH, "(//input[@id='ctl00_body_mlkp_-3_lookupControl_lookupText'])[1]")
            input_plan.send_keys(plan)

            input_character = browser.find_element(By.XPATH, "(//input[@id='ctl00_body_mslkp_особенность_lookupControl_lookupText'])[1]")
            input_character.send_keys(character)

            # Отправляем форму
            button = browser.find_element(By.XPATH, "(//span[@id='ctl00_body_ctl27'])[1]")
            button.click()

            sheet.cell(row=i + 2, column=5).value = "успешно"

        except:

            sheet.cell(row=i + 2, column=5).value = "ошибка"

        workbook.save('RDBTASKSFIND.xlsx')

finally:

    browser.quit()
    workbook.close()